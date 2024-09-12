import zipfile
import os
import re
import json
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import NamedStyle, NumberFormatDescriptor, Font

def process_data(json_file_path):
    with open(json_file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)

    result = []
    for message in data["messages"]:
        if message.get("from") != "Tribute":
            continue

        sender_name = extract_sender_name(message.get("text_entities", []))
        message_date = datetime.fromisoformat(message.get("date"))
        payment_type, amount = extract_payment_info(message.get("text_entities", []))
        
        if payment_type and amount is not None:
            result.append({
                "Дата и время": message_date,
                "Пользователь": sender_name,
                "Сумма": amount,
                "Категория": payment_type
            })

    df = create_dataframe_with_quarters(result)
    return df

def extract_sender_name(text_entities):
    for entity in text_entities:
        if entity.get("type") == "mention":
            return entity.get("text", "Неизвестно")
        elif entity.get("type") == "mention_name":
            name = entity.get("text", "Неизвестно")
            user_id = entity.get("user_id", "")
            return f"{name} (id{user_id})"
    return "Неизвестно"

def extract_payment_info(text_entities):
    # Инициализация переменных
    payment_type = None
    amount = None

    # Объединение всех текстовых сущностей в одну строку для определения типа платежа
    full_text = ' '.join([entity.get("text", "").lower() for entity in text_entities])

    # Определение типа платежа
    if "новая подписка" in full_text or "оформил подписку" in full_text:
        payment_type = "Новая подписка"
    elif "продлена подписка" in full_text or "продлил подписку" in full_text:
        payment_type = "Обновление подписки"
    elif "новый донат" in full_text or "отправил донат" in full_text:
        payment_type = "Донат"

    # Поиск суммы платежа
    for entity in text_entities:
        if entity.get("type") == "bold":
            match = re.search(r"[₽]\d{1,7}\.\d{2}", entity.get("text", ""))
            if match:
                # Преобразование формата суммы для корректного отображения в Excel
                amount = float(match.group().replace('₽', '').replace('€', '').replace(',', '.'))
                break

    return payment_type, amount

def create_dataframe_with_quarters(data):
    df = pd.DataFrame(data)

    # Функция для определения квартала
    def get_quarter(month):
        if 1 <= month <= 3:
            return 'Q1'
        elif 4 <= month <= 6:
            return 'Q2'
        elif 7 <= month <= 9:
            return 'Q3'
        else:
            return 'Q4'

    # Добавление столбца с кварталами
    df['Квартал'] = df['Дата и время'].apply(lambda x: get_quarter(x.month))
    return df

def style_excel_sheet(worksheet, start_row):
    # Настройки форматирования
    text_format = '@'  # Текстовый формат
    date_style = NamedStyle(name='datetime', number_format='DD.MM.YYYY HH:MM')
    currency_style = NamedStyle(name='currency', number_format='# ##0.00 ₽')

    # Форматирование основной таблицы, начиная с start_row
    for row in worksheet.iter_rows(min_row=start_row, max_col=4, max_row=worksheet.max_row):
        # Форматирование столбца с датой
        row[0].style = date_style

        # Форматирование столбцов "Пользователь" и "Категория" как текст
        for cell in [row[1], row[3]]:
            cell.number_format = text_format

        # Форматирование столбца с суммой
        row[2].style = currency_style

    # Установка ширины столбцов
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 22
    worksheet.column_dimensions['C'].width = 14
    worksheet.column_dimensions['D'].width = 22

    # Установка выравнивания текста по левому краю
    for col in worksheet.iter_cols(min_row=start_row, max_col=4, max_row=worksheet.max_row):
        for cell in col:
            cell.alignment = openpyxl.styles.Alignment(horizontal='left')

def calculate_summary(df):
    # Инициализация словаря для сводной информации
    summary = {'Донат': [0, 0.0], 'Новые и обновлённые подписки': [0, 0.0], 'Налог 6% с подписок': [0, 0.0], 'Всего платежей': [0, 0.0]}
    
    # Расчет общего количества и суммы для Донатов
    if 'Донат' in df['Категория'].values:
        donat_summary = df[df['Категория'] == 'Донат']['Сумма'].agg(['count', 'sum'])
        summary['Донат'] = [donat_summary['count'], donat_summary['sum']]

    # Расчет общего количества и суммы для Новых и Обновленных подписок
    subscription_df = df[df['Категория'].isin(['Новая подписка', 'Обновление подписки'])]
    if not subscription_df.empty:
        subscription_summary = subscription_df['Сумма'].agg(['count', 'sum'])
        summary['Новые и обновлённые подписки'] = [subscription_summary['count'], subscription_summary['sum']]

        # Расчет налога с подписок
        summary['Налог 6% с подписок'] = [0, subscription_summary['sum'] * 0.06]

    # Общее количество и сумма платежей
    summary['Всего платежей'] = [df.shape[0], df['Сумма'].sum()]

    return summary

def add_summary_to_excel(worksheet, summary):
    # Вставка сводной информации перед основной таблицей
    for r_idx, (key, values) in enumerate(summary.items(), 1):
        # Вставка названия категории
        worksheet.cell(row=r_idx, column=1, value=key)

        # Вставка количества и суммы
        for c_idx, value in enumerate(values, 2):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            if c_idx == 3:  # Форматирование столбца с суммой
                cell.number_format = '# ##0.00 ₽'

def save_to_excel(df):
    output_files = []
    quarter_summaries = {}
    quarter_to_months = {
        'Q1': '_янв_фев_мар',
        'Q2': '_апр_май_июн',
        'Q3': '_июл_авг_сен',
        'Q4': '_окт_ноя_дек'
    }

    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        quarter_df = df[df['Квартал'] == quarter]
        if quarter_df.empty:
            quarter_summaries[quarter + quarter_to_months[quarter]] = "За квартал {} данных нет".format(quarter)
            continue

        quarter_df = quarter_df.drop(columns=['Квартал'])
        summary = calculate_summary(quarter_df)

        # Форматирование сводных данных для возврата
        formatted_summary = '\n'.join([
            f"{key} - {int(value[0])}шт, {value[1]:,.2f}₽" if key != 'Налог 6% с подписок' else f"{key} - {value[1]:,.2f}₽"
            for key, value in summary.items()
        ])
        quarter_summaries[quarter + quarter_to_months[quarter]] = formatted_summary

        # Измененный формат имени файла
        file_name = f'/mnt/data/Квартал_{quarter}{quarter_to_months[quarter]}.xlsx'
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            workbook = writer.book
            worksheet = workbook.create_sheet(f'Квартал {quarter}{quarter_to_months[quarter]}')

            add_summary_to_excel(worksheet, summary)

            # Оставляем одну пустую строку между сводной информацией и основной таблицей
            df_start_row = worksheet.max_row + 2

            # Сохранение данных
            for r_idx, row in enumerate(quarter_df.to_numpy(), df_start_row):
                for c_idx, value in enumerate(row, 1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)

            style_excel_sheet(worksheet, df_start_row)

        output_files.append(file_name)

    return output_files, quarter_summaries